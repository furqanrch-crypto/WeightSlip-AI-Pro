from __future__ import annotations

from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DUPLICATE_FILL = PatternFill("solid", fgColor="FFF2CC")
DUPLICATE_SLIP_FILL = PatternFill("solid", fgColor="F4B183")
REVIEW_FILL = PatternFill("solid", fgColor="FCE4D6")
SUMMARY_FILL = PatternFill("solid", fgColor="D9EAF7")


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
        ws.column_dimensions[column_letter].width = min(max(length + 2, 12), 42)


def _style_header(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_excel_report(records: Iterable, output_path: Path) -> Path:
    records = list(records)
    wb = Workbook()
    ws = wb.active
    ws.title = "Weight Slips"

    headers = [
        "Sr No",
        "Slip No",
        "Vehicle",
        "Party",
        "Product",
        "1st Weight (Kg)",
        "2nd Weight (Kg)",
        "Net Weight (Kg)",
        "1st Date/Time",
        "2nd Date/Time",
        "Location",
        "Operator",
        "Status",
        "Validation",
        "Duplicate",
        "Duplicate Of",
        "OCR Confidence",
        "Original File",
    ]
    ws.append(headers)
    _style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:R{max(1, len(records) + 1)}"

    slip_counts = Counter(str(r.slip_no).strip() for r in records if r.slip_no)

    for index, record in enumerate(records, start=1):
        row = [
            index,
            record.slip_no,
            record.vehicle_no,
            record.party,
            record.product,
            record.first_weight,
            record.second_weight,
            record.net_weight,
            record.first_datetime,
            record.second_datetime,
            record.location,
            record.operator,
            record.processing_status,
            record.validation_status,
            "YES" if record.duplicate else "NO",
            record.duplicate_of,
            record.confidence,
            record.original_filename,
        ]
        ws.append(row)
        excel_row = ws.max_row

        repeated_number = bool(record.slip_no and slip_counts[str(record.slip_no).strip()] > 1)
        if repeated_number or record.duplicate:
            for cell in ws[excel_row]:
                cell.fill = DUPLICATE_FILL
            ws.cell(excel_row, 2).fill = DUPLICATE_SLIP_FILL
            ws.cell(excel_row, 2).font = Font(bold=True)
        elif record.processing_status in {"review_required", "failed"}:
            for cell in ws[excel_row]:
                cell.fill = REVIEW_FILL

        for col in (6, 7, 8):
            ws.cell(excel_row, col).number_format = '#,##0.00'
        ws.cell(excel_row, 17).number_format = '0.00%'

    _autosize(ws)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "WeightSlip AI Pro - Summary"
    summary["A1"].font = Font(size=16, bold=True)
    summary.merge_cells("A1:D1")
    summary["A2"] = "Generated"
    summary["B2"] = datetime.now().strftime("%d-%b-%Y %I:%M %p")

    total_slips = len(records)
    unique_slips = len({str(r.slip_no).strip() for r in records if r.slip_no})
    duplicate_rows = sum(1 for r in records if r.duplicate)
    repeated_slip_numbers = sum(1 for count in slip_counts.values() if count > 1)
    completed = sum(1 for r in records if r.processing_status in {"completed", "duplicate"})
    review_required = sum(1 for r in records if r.processing_status == "review_required")
    failed = sum(1 for r in records if r.processing_status == "failed")

    # Business totals exclude duplicate rows so the same weighbridge slip is not counted twice.
    original_records = [r for r in records if not r.duplicate]
    total_first = sum(float(r.first_weight or 0) for r in original_records)
    total_second = sum(float(r.second_weight or 0) for r in original_records)
    total_net = sum(float(r.net_weight or 0) for r in original_records)

    metrics = [
        ("Total Images / Records", total_slips),
        ("Unique Slip Numbers", unique_slips),
        ("Duplicate Rows", duplicate_rows),
        ("Repeated Slip Numbers", repeated_slip_numbers),
        ("OCR Completed", completed),
        ("Review Required", review_required),
        ("Failed", failed),
        ("Total 1st Weight (Kg) - excluding duplicates", total_first),
        ("Total 2nd Weight (Kg) - excluding duplicates", total_second),
        ("Total Net Weight (Kg) - excluding duplicates", total_net),
        ("Total Net Weight (MT) - excluding duplicates", total_net / 1000 if total_net else 0),
    ]

    summary.append([])
    summary.append(["Metric", "Value"])
    _style_header(summary, summary.max_row)
    for metric, value in metrics:
        summary.append([metric, value])
        summary.cell(summary.max_row, 1).fill = SUMMARY_FILL
        if isinstance(value, float):
            summary.cell(summary.max_row, 2).number_format = '#,##0.00'

    product_summary = defaultdict(lambda: {"count": 0, "net": 0.0})
    for record in original_records:
        product = record.product or "Unknown"
        product_summary[product]["count"] += 1
        product_summary[product]["net"] += float(record.net_weight or 0)

    summary.append([])
    summary.append(["Product", "Slip Count", "Net Weight (Kg)", "Net Weight (MT)"])
    _style_header(summary, summary.max_row)
    for product, values in sorted(product_summary.items()):
        summary.append([
            product,
            values["count"],
            values["net"],
            values["net"] / 1000 if values["net"] else 0,
        ])
        summary.cell(summary.max_row, 3).number_format = '#,##0.00'
        summary.cell(summary.max_row, 4).number_format = '#,##0.000'

    duplicate_sheet = wb.create_sheet("Duplicates")
    duplicate_sheet.append([
        "Slip No", "Occurrences", "Record IDs", "Vehicles", "Parties", "Net Weights (Kg)"
    ])
    _style_header(duplicate_sheet)

    grouped = defaultdict(list)
    for record in records:
        if record.slip_no:
            grouped[str(record.slip_no).strip()].append(record)

    for slip_no, group in sorted(grouped.items()):
        if len(group) <= 1:
            continue
        duplicate_sheet.append([
            slip_no,
            len(group),
            ", ".join(str(r.id) for r in group),
            ", ".join(sorted({r.vehicle_no for r in group if r.vehicle_no})),
            ", ".join(sorted({r.party for r in group if r.party})),
            ", ".join(f"{float(r.net_weight):,.0f}" for r in group if r.net_weight is not None),
        ])
        for cell in duplicate_sheet[duplicate_sheet.max_row]:
            cell.fill = DUPLICATE_FILL
        duplicate_sheet.cell(duplicate_sheet.max_row, 1).fill = DUPLICATE_SLIP_FILL
        duplicate_sheet.cell(duplicate_sheet.max_row, 1).font = Font(bold=True)

    _autosize(summary)
    _autosize(duplicate_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
